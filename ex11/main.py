import openpyxl 

book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet('fruits')

fruits_page = book['fruits']

fruits_page.append(['Fruits', 'Quantity', 'Price'])
fruits_page.append(['banana', '5', 'R$4,90'])
fruits_page.append(['fruit 2', '2', 'R$6,99'])
fruits_page.append(['fruit 3', '12', 'R$12,29'])
fruits_page.append(['fruit 4', '20', 'R$3,90'])

book.save('shopping-spreadsheet.xlsx')