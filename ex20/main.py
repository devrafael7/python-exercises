from flask import Flask, request, jsonify, render_template
import openpyxl
import os
import random
from datetime import datetime, timedelta

app = Flask(__name__)

file_name = 'test_py03.xlsx'

if os.path.exists(file_name):
    book = openpyxl.load_workbook(file_name)
else:
    book = openpyxl.Workbook()
    book.create_sheet('Info')

info_page = book['Info']

if 'Info' not in book.sheetnames or book['Info'].max_row == 1:
      info_page.append(['Product Name', 'Product ID', 'Price', 'Place', 'Quantity', 'Date', 'Time', 'Order ID'])
    

unique_id = set()
unique_order_id = set()

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/get_data', methods=['POST'])
def receber_data():
    data = request.get_json()

    data_name = data.get('dataName')
    data_price = data.get('dataPrice')
    data_place = data.get('dataPlace')
    data_quantity = data.get('dataQuantity')
    
    while True:
        product_id = random.randint(10000,99999)
        if product_id not in unique_id:
            unique_id.add(product_id)
        break
    
    while True:
        order_id = f"ORD-{random.randint(100000,999999)}"
        if order_id not in unique_order_id:
            unique_order_id.add(order_id)
        break
    
    date_now = datetime.now().strftime('%Y-%m-%d')
    time_now = datetime.now().strftime('%H:%M:%S')
    

    print(f'Data-name recebido: {data_name} {data_price} {data_place} {data_quantity}')
    
    info_page.append([data_name, product_id, data_price, data_place, data_quantity, date_now, time_now, order_id])
    
    book.save('test_py03.xlsx')

    return jsonify({'message': f'Recebido: {data_name, product_id, data_price, data_place, data_quantity, date_now, time_now, order_id}'})


if __name__ == '__main__':
    app.run(debug=True)