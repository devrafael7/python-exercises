import openpyxl
import tkinter as tk

arquivo_excel = openpyxl.load_workbook('Estoque (1).xlsx')
pagina = arquivo_excel.active

primeira_linha = []
for cell in pagina[1]:
    primeira_linha.append(cell.value)
    

produto = input('digite o produto: ')
linhas_produto = []
for cada_linha in pagina.iter_rows(min_row=2, values_only=True):
    linha_encontrada = False
    for cell in cada_linha:
        if produto in str(cell):
            linha_encontrada = True
            break
    
    if linha_encontrada:
        linhas_produto.append(cada_linha)

estoque_total = 0

for quantidade in linhas_produto:
    estoque_total += quantidade[4]

print(primeira_linha)
print(linhas_produto)
print(estoque_total)



























##import openpyxl
##import tkinter as tk
##
##book = openpyxl.load_workbook('Estoque (1).xlsx')
##main_page = book.active
##
##dados = input('digite algo: ')
##
##
##first_row = []
##for cell in main_page[1]:
##    first_row.append(cell.value)
##    
##todas_linhas = []
##
##for cada_linha in main_page.iter_rows(min_row=2, values_only=True):
##    linha_encontrada = False
##    for cell in cada_linha:
##        if dados in str(cell):
##            linha_encontrada = True
##            break
##        
##    if linha_encontrada:
##        todas_linhas.append(cada_linha)
##        
##print(first_row)
##print(todas_linhas)
##

##import openpyxl
##import tkinter as tk
##
### Carrega a planilha e ativa a primeira aba
##book = openpyxl.load_workbook('Estoque (1).xlsx')
##main_sheet = book.active
##
### Cria uma lista com os valores da primeira linha (cabeçalho)
##first_row = [cell.value for cell in main_sheet[1]]
##
### Solicita a entrada do usuário
##dados = input('Digite algo para buscar: ')
##
### Lista para armazenar as linhas que contenham o dado buscado
##info = []
##
### Percorre as linhas da planilha a partir da segunda linha (min_row=2)
##for linha in main_sheet.iter_rows(min_row=2, values_only=True):
##    encontrado = False  # Define uma flag para verificar se encontrou o dado na linha
##    for cell in linha:
##        if dados in str(cell):  # Verifica se o dado digitado está na célula
##            encontrado = True   # Se encontrado, define a flag como True
##            break               # Sai do loop interno assim que encontrar o dado
##    if encontrado:  # Verifica se a flag foi marcada como True após o loop
##        info.append(linha)  # Adiciona a linha inteira à lista `info`
##        
##
##root = tk.Tk()
##root.title('teste')
##root.geometry('400x400')
##        
##info_text = tk.Text(root, font=('Arial', 12))
##info_text.pack( fill='both')
##for linha in info:
##    info_text.insert('end', str(linha))
##
### Exibe o resultado
##print(info)
##
##root.mainloop()
        
    