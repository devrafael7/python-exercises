from openpyxl import load_workbook

def refresh_file(path_origin, path_destination):
    
    origin_wb = load_workbook(path_origin)
    destination_wb = load_workbook(path_destination)
    
    origin_sheet = origin_wb.active
    destination_sheet = destination_wb.active
    
    existent_data = set()
    for row in destination_sheet.iter_rows(values_only=True):
        existent_data.add(tuple(row))
        
    new_data_added = False
    for row in origin_sheet.iter_rows(values_only=True):
        if tuple(row) not in existent_data:
            destination_sheet.append(row)
            new_data_added = True
    
    if new_data_added:
        destination_wb.save(path_destination)
        print("New data were added successfully!")
    else:
        print("No data found to add in!")
        
    origin_wb.close()
    destination_wb.close()
    

origin_path = r"C:\Users\Rafar\OneDrive\Área de Trabalho\base1.xlsx"
destination_path = r"C:\Users\Rafar\OneDrive\Área de Trabalho\base2.xlsx"

refresh_file(origin_path, destination_path)


        
    
