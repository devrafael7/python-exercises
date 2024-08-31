from flask import Flask, request, render_template
import openpyxl
from datetime import datetime
import os

app = Flask(__name__)

xlsx_file = 'registers.xlsx'


if os.path.exists(xlsx_file):
    book = openpyxl.load_workbook(xlsx_file)
    if 'already_account' in book.sheetnames:
        already_account_page = book['already_account']
    else:
        already_account_page = book.create_sheet('already_account')
        already_account_page.append(['Nickname', 'Password'])
        
    if 'register' in book.sheetnames:
        register_page = book['register']
    else:
        register_page = book.create_sheet('register')

        register_page.append(['Name', 'E-mail', 'Password', 'Creation Date', 'Creation Time'])
else:
    book = openpyxl.Workbook()
    register_page = book.active
    register_page.title = 'register'
    register_page.append(['Name', 'E-mail', 'Password', 'Creation Date', 'Creation Time'])
    
    already_account_page = book.active
    already_account_page.title = 'already_account'
    register_page.append(['Name', 'E-mail', 'Password', 'Creation Date', 'Creation Time'])
    
    book.save(xlsx_file)

@app.route('/', methods=['GET', 'POST'])
def index_register():
    if request.method == 'POST':
        register_name = request.form['register_name']
        register_email = request.form['register_email']
        register_password = request.form['register_password']
        register_date = datetime.now().strftime('%Y-%m-%d')
        register_time = datetime.now().strftime('%H:%M:%S')

        email_exists = False
        for row in register_page.iter_rows(values_only=True):
            if row[1] == register_email:
                email_exists = True
                break
        
        if email_exists:
            return render_template('already_account.html')
        else:
   
            register_page.append([register_name, register_email, register_password, register_date, register_time])

            book.save(xlsx_file)
            return render_template('create_user.html')

    return render_template('index.html')

app.route('/', methods=['GET','POST'])
def already_account():
    already_nickname = request.form['already_nickname']
    already_password = request.form['already_password']
    
    nickname_exists = False
    for row in already_account_page.iter_rows(values_only=True):
        if row[1] == already_nickname:
            nickname_exists = True
            break
    
    if nickname_exists:
        return render_template('main.html')
    else:
        return render_template('index.html')
    
app.route('/', methods=['GET', 'POST'])
def create_user():
    create_nickname = request.form['create_nickname']
    create_password = request.form['create_password']

if __name__ == '__main__':
    app.run(debug=True)