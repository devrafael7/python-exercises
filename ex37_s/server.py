from flask import Flask, request, render_template
from datetime import datetime
import openpyxl

app = Flask(__name__)

register_db = 'register_db.xlsx'

# Carregar o banco de dados de registros
book = openpyxl.load_workbook(register_db)
register_page = book.active

@app.route('/', methods=['GET', 'POST'])
def sign_up():
    show_element = False  # Inicialmente, o elemento está oculto
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        register_date = datetime.now().strftime('%Y-%m-%d')
        register_time = datetime.now().strftime('%H:%M:%S')
        
        # Verificar se o email já existe
        email_exists = any(row[0] == email for row in register_page.iter_rows(values_only=True))
        
        if email_exists:
            show_element = True  # Se o email já existir, o elemento será mostrado
            return render_template('sign_up.html', email_exists_on_db=True, show_element=show_element)
        else:
            # Adicionar o novo registro
            register_page.append([email, password, register_date, register_time])
            book.save(register_db)
            return render_template('index.html')
        
    return render_template('sign_up.html', email_exists_on_db=False, show_element=show_element)


@app.route('/sign_in', methods=['GET', 'POST'])
def sign_in():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        valid_user = False
        
        for row in register_page.iter_rows(values_only=True):
            if row[0] == email and row[1] == password:
                valid_user = True
                break
        
        if valid_user:
            return render_template('index.html')
        else:
            return render_template('sign_in.html', login_failed=True)  # Exemplo de variável para erro
        
    return render_template('sign_in.html')  # Corrigido: adicionado return

if __name__ == '__main__':
    app.run(debug=True)
