from flask import Flask, request, render_template
import openpyxl
from datetime import datetime
import os

app = Flask(__name__)

excel_file = 'flask.xlsx'

if os.path.exists(excel_file):
    book = openpyxl.load_workbook(excel_file)
    if 'register' in book.sheetnames:
        register_page = book['register']
    else:
        register_page = book.create_sheet('register')
        register_page.append(['Nickname', 'E-mail', 'Password', 'Date of Creation', 'Time of Creation'])
else:
    book = openpyxl.Workbook()
    register_page = book.active
    register_page.title = 'register'
    register_page.append(['Nickname', 'E-mail', 'Password', 'Date of Creation', 'Time of Creation'])
    book.save(excel_file)

@app.route('/', methods=['GET', 'POST'])
def sign_up():
    if request.method == 'POST':
        nickname = request.form['nickname']
        email = request.form['email']
        password = request.form['password']
        register_date = datetime.now().strftime('%Y-%m-%d')
        register_time = datetime.now().strftime('%H:%M:%S')
        
        nickname_exists = False
        email_exists = False
        for row in register_page.iter_rows(values_only=True):
            if row[0] == nickname:
                nickname_exists = True
            if row[1] == email:
                email_exists = True
            
        if nickname_exists or email_exists:
            return render_template('sign_in.html')
        else:
            register_page.append([nickname, email, password, register_date, register_time])
            book.save(excel_file)
            return render_template('index.html')
        
    return render_template('sign_up.html')

@app.route('/sign_in', methods=['GET', 'POST'])
def sign_in():
    if request.method == 'POST':
        nickname = request.form['nickname']
        password = request.form['password']
    
        valid_user = False
        
        for row in register_page.iter_rows(values_only=True):
            if row[0] == nickname and row[2] == password:
                valid_user = True
                break
        
        if valid_user:
            return render_template('index.html')
        else:
            return render_template('sign_in.html')
    return render_template('sign_in.html')

if __name__ == '__main__':
    app.run(debug=True)