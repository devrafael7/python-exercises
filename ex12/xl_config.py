import openpyxl

book = openpyxl.load_workbook('objects-info.xlsx')

objectsPage = book['Objects']

totalPrice = 0

for rows in objectsPage.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in rows:
        totalPrice += float(cell.value)

for xlRows in objectsPage.iter_rows(min_row=1, max_row=7):
    for eachCell in xlRows:
        if eachCell.value == 'Fork':
            eachCell.value == 'Spoon'
        ##if eachCell is not None:
        ##    print(eachCell.value)
                
book.save('objects-info.xlsx')