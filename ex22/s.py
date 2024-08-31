import socket
import openpyxl
from fpdf import FPDF

book = openpyxl.load_workbook('info.xlsx')
info_page = book['Info']

def buscar_linhas_celular(sheet):
    linhas_celular = []
    
    primeira_linha = [str(cell.value) for cell in sheet[1]]
    
    linhas_celular.append(primeira_linha)
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any('Laptop' in str(cell) for cell in row if cell):
            linhas_celular.append(row)
    
    return linhas_celular

server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

server_socket.bind(('localhost', 12345))

server_socket.listen(1)

print("Servidor aguardando conexão...")

client_socket, client_address = server_socket.accept()

print(f"Conexão estabelecida com {client_address}")

data = client_socket.recv(1024).decode()

if data == 'Laptop':
    linhas_celular = buscar_linhas_celular(info_page)

    pdf = FPDF()

    pdf.add_page()

    pdf.set_font("Arial", size=8)

    for linha in linhas_celular:
        linha_texto = ', '.join([str(cell) for cell in linha])
        pdf.cell(200, 10, txt=linha_texto, ln=True, align='C')

    pdf.output("info.pdf")

    print("PDF gerado com sucesso!")
else:
    print('Nenhuma ação realizada. Tente novamente.')

client_socket.send("Mensagem recebida pelo servidor!".encode())

client_socket.close()
server_socket.close()