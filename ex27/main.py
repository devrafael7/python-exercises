import openpyxl

book = openpyxl.load_workbook('Estoque (1).xlsx')
first_page = book.active

first_row = []

for cell in first_page[1]:
    first_row.append(cell.value)
    

dados = input('type the name of product: ')

all_products = []

for each_row in first_page.iter_rows(min_row=2, values_only=True):
    product_found = False
    
    for cell in each_row:
        if dados in str(cell):
            product_found = True
            break
    
    if product_found:
        all_products.append(each_row)

total_stock = 0

for quantity in all_products:
    total_stock += quantity[4]

print(first_row)
print(all_products)
print(total_stock)

##import openpyxl
##
##arquivo_excel = openpyxl.load_workbook('Estoque (1).xlsx')
##pagina = arquivo_excel.active
##
##primeira_linha = []
##
##for cell in pagina[1]:
##    primeira_linha.append(cell.value)
##
##
##dados = input('digite a marca do produto: ')
##
##todos_produtos = []
##
##for cada_linha in pagina.iter_rows(min_row=2, values_only=True):
##    produto_encontrado = False
##    
##    for cell in cada_linha:
##        if dados in str(cell):
##            produto_encontrado = True
##            break
##    
##    if produto_encontrado:
##        todos_produtos.append(cada_linha)
##
##estoque_total = 0
##
##for quantidade in todos_produtos:
##    estoque_total += quantidade[4]
##
##print(primeira_linha)
##print(todos_produtos)
##print(estoque_total)

##import openpyxl
##
##arquivo_excel = openpyxl.load_workbook('Estoque (1).xlsx')
##
##pagina = arquivo_excel.active
##
##primeira_linha = []
##
##for cell in pagina[1]:
##    primeira_linha.append(cell.value)
##
##dados = input('digite o nome do produto: ')
##
##produtos = []
##
##for cada_linha in pagina.iter_rows(min_row=2, values_only=True):
##    produto_encontrado = False
##    for cell in cada_linha:
##        if dados in str(cell):
##            produto_encontrado = True
##            break
##    
##    if produto_encontrado:
##        produtos.append(cada_linha)
##
##estoque_total = 0
##for quantidade in produtos:
##    estoque_total += quantidade[4]
##    
##print(primeira_linha)
##print(produtos)
##print(estoque_total)

##import openpyxl
##import tkinter as tk
##
##arquivo_excel = openpyxl.load_workbook('Estoque (1).xlsx')
##primeira_pagina = arquivo_excel.active
##
##primeira_linha = []
##
##for cell in primeira_pagina[1]:
##    primeira_linha.append(cell.value)
##
##
##nome_produto = input('digite o produto: ')
##
##produtos_encontrados = []
##
##for linha in primeira_pagina.iter_rows(min_row=2, values_only=True):
##    linha_encontrada = False
##    for cell in linha:
##        if nome_produto in str(cell):
##            linha_encontrada = True
##            break
##    
##    if linha_encontrada:
##        produtos_encontrados.append(linha)
##
##estoque_total = 0
##
##for cell in produtos_encontrados:
##    estoque_total += cell[4]
##    
##
##print(primeira_linha)
##print(produtos_encontrados)
##print(estoque_total)
##
##root = tk.Tk()
##root.title('sei nao')
##root.geometry('400x400')
##
##app_text = tk.Text(root, font=('Arial', 10))
##app_text.pack(fill='both')
##
##app_text.insert('end', str(produtos_encontrados))
##
##root.mainloop()