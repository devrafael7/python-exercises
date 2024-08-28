from flask import Flask, request, render_template
import openpyxl
import os
from datetime import datetime
import random

app = Flask(__name__)

excel_file = 'form_py.xlsx'

if os.path.exists(excel_file):
    book = openpyxl.load_workbook(excel_file)
else:
    book = openpyxl.Workbook()
    book.create_sheet('Form')
    
    
form_page = book['Form']

if 'Form' not in book.sheetnames or book['Form'].max_row == 1:
    form_page.append(['Purchaser Name', 'Gender', 'Location', 'Product', 'Product ID', 'Price', 'Quantity', 'Date', 'Time', 'Order ID'])


@app.route('/', methods=['GET', 'POST'])
def form():
    if request.method == 'POST':
        name = request.form['name_value']
        gender = request.form['gender_type']
        location = request.form['location']
        product = request.form['product_name']
        price = request.form['price_value']
        quantity = request.form['quantity']
        
        date = datetime.now().strftime('%Y-%m-%d')
        time = datetime.now().strftime('%H:%M:%S')
        
        unique_PI = set()
        
        while True:
            product_id = random.randint(1, 10000)
            if product_id not in unique_PI:
                unique_PI.add(product_id)
            break
        
        unique_OD = set()
        
        while True:
            order_id = f"ORD-{random.randint(10000, 99999)}"
            if order_id not in unique_OD:
                unique_OD.add(order_id)
            break
        
        form_page.append([name, gender, location, product, product_id, price, quantity, date, time, order_id])
        book.save('form_py.xlsx')

        return 'deu bom mermo'
    return render_template('index.html')
        
if __name__ == '__main__':
    app.run(debug=True)


